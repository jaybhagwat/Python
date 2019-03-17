#HW:Read optparse module and write the argument parser program using optparse.

#Write a program to read an excel file and convert it into corresponding vcf.
import openpyxl

def XmlToVcf(input_file):
	fd=openpyxl.load_workbook(input_file)
	wb=fd.active
	#names=wb.sheetnames()
	#print(names)
	print(wb.max_column)
	print(wb.max_row)
	str_start="BEGIN : VCARD" + " \n" + "VERSION :2.1"+"\n"
	str_end="VCARD"
	vcard=open("Vcard.vcf","w")
	
	for i in range(1,wb.max_row+1):
		fn=wb.cell(row=i,column=1).value
		org=wb.cell(row=i,column=4).value
		tel=wb.cell(row=i,column=3).value
		email=wb.cell(row=i,column=2).value
		vcard.write(str_start)
		vcard.write("FN:"+fn+"\n")
		vcard.write("ORG:"+org+"\n")
		vcard.write("TEL:"+str(tel)+"\n")		
		vcard.write("EMAIL:"+email+"\n")
		vcard.write(str_end)
		
	fd.close()
	vcard.close()

def main():
	input_file=eval(input("Enter the file name"))
	XmlToVcf(input_file)

if __name__=='__main__':
	main()
	
	
#getsheetnames
#getsheetby names	